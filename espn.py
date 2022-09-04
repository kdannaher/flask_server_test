# ---------------------- Imports ----------------------------
from bs4 import BeautifulSoup
import time
from selenium import webdriver
import string
import random
import openpyxl
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.common.keys import Keys


#-----------------------Settings--------------------------
username = 'DoeJohn41206@gmail.com'
password = 'JohnDoe41206'
domain = "https://fantasy.espn.com/football/livedraftresults?leagueId=38910595"

class EspnScrapper(object):
    def __init__(self, username, password, domain) -> None:
        # initialize the browser here
        self.username =  username
        self.password =  password
        self.domain = domain
        print(self.username, self.password)
        self.error_message = ''
        self.success_response = ''
        co = webdriver.ChromeOptions()
        #co.add_argument("--headless") # removes the browser gui
        co.add_argument("--log-level=3") # makes the log less talkative
        co.add_experimental_option("detach", True)
        self.driver = webdriver.Chrome('.\chromedriver.exe', options = co)
        self.driver.get(self.domain)
        print("passed 4")
        print("Fantasy draft loaded at", self.driver.current_url)

    def random_string_file_name(self):
        """
        Generate random_filename_for the string
        """
        return ''.join(random.choices(string.ascii_letters, k=9))
            
    def cookie_accepter(self):
        """
        This checks for the cookies and check if the tab is available.
        """
        try:
            print("Waiting for cookie dialog...")
            WebDriverWait(self.driver, 20).until(EC.presence_of_element_located(('id', 'onetrust-accept-btn-handler')))
            self.driver.find_element('id', 'onetrust-accept-btn-handler').click()
            print("Accepted Cookies")
        except TimeoutException:
            print("No Request For Cookies. Proceeding To Login...")
    
    def login_website(self):
        """
        This method is used to login to the website
        """
        try:
            print("Waiting for login...")
            self.driver.refresh()
            WebDriverWait(self.driver, 25).until(EC.presence_of_element_located(('xpath', '/html/body/div[2]/iframe')))
            the_iframe = self.driver.find_element('xpath',"/html/body/div[2]/iframe")
            self.driver.switch_to.frame(the_iframe)
            user = WebDriverWait(self.driver, 25).until(EC.presence_of_element_located(('xpath', '//*[@id="did-ui-view"]/div/section/section/form/section/div[1]/div/label/span[2]/input')))
            user.send_keys(self.username)
            password_input = WebDriverWait(self.driver, 25).until(EC.presence_of_element_located(('xpath', '//*[@id="did-ui-view"]/div/section/section/form/section/div[2]/div/label/span[2]/input')))
            password_input.send_keys(self.password)
            time.sleep(0.5)
            password_input.send_keys(Keys.RETURN)
            time.sleep(0.5)
            self.driver.switch_to.default_content()
            WebDriverWait(self.driver, 25).until(EC.presence_of_element_located(('class name', 'Table')))
            print("Logged In")
            return "Successfully logged in", 0
        except TimeoutError:
            print("Draft page didn't load.")
            # self.driver.quit()
            return "login_request_failed", -1
    
    def create_excel_template(self):
        # ---------------------- Excel Setup ----------------------------
        wb = openpyxl.Workbook()
        ws = wb.active
        ws['A1'] = "RANK"
        ws['B1'] = "PLAYER"
        ws['C1'] = "TEAM"
        ws['D1'] = "POSITION"
        ws['E1'] = "SNAKE DRAFT AVG PICK"
        ws['F1'] = "SNAKE DRAFT 7 DAY +/-"
        ws['G1'] = "AVG SALARY"
        ws['H1'] = "SALARY 7 DAY +/-"
        ws['I1'] = "%ROST"
        print("Draft Sheet Prepared, scraping begins")
        return ws, wb
    
    def main_loop(self):
        # ---------------------- Main Loop ----------------------------
        self.cookie_accepter()
        
        return_message, status_code = self.login_website()
        if status_code == - 1:
            self.error_message += return_message
            return {"status_code": -1, "error_message": self.error_message, "success_message": self.success_response}
        else:
            self.success_response += return_message

        ws,wb = self.create_excel_template()
        # Start the scrapping from here
        # goes through every page of player listings
        file_name = self.random_string_file_name()
        invalid_button = self.driver.find_element('class name', "Button--disabled")
        while self.driver.find_elements('class name', 'Button')[1] != invalid_button:
            # Check if we've reached the end of the list
            try: 
                invalid_button = self.driver.find_element('class name', "Button--disabled") # only valid at the end
            except: 
                invalid_button = None # keeps the loop going on the next turn

            soup = BeautifulSoup(self.driver.page_source, 'lxml')
            table = soup.find('table')
            table_rows = table.tbody.find_all('tr')

            for tr in table_rows: # iterate through every row in the table
                arr = [] # very useful array, for temp table storage
                for j, td in enumerate(tr.find_all('td')): # " " every cell in the row
                    if j == 1: # player name entry split into name, team, and position
                        arr+=[td.div['title']]
                        team = td.div.find('span', {'class':'playerinfo__playerteam'}).getText()
                        pos = td.div.find('span', {'class':'playerinfo__playerpos ttu'}).getText()
                    else:
                        arr+=[td.getText()]

                # Starts populating the excel file
                ws['A'+str(int(arr[0])+1)] = arr[0]
                ws['B'+str(int(arr[0])+1)] = arr[1]
                ws['C'+str(int(arr[0])+1)] = team
                ws['D'+str(int(arr[0])+1)] = pos
                ws['E'+str(int(arr[0])+1)] = arr[2]
                ws['F'+str(int(arr[0])+1)] = arr[3]
                ws['G'+str(int(arr[0])+1)] = arr[4]
                ws['H'+str(int(arr[0])+1)] = arr[5]
                ws['I'+str(int(arr[0])+1)] = arr[6]

            # saves the workbook when done with a page
            wb.save(f"./excel_files/{file_name}.xlsx")

            # Goes to the next page
            self.driver.find_elements('class name', 'Button')[1].click()
            old_table_rows = table_rows
            v = 1

            # checks if next page has loaded and waits till it does
            while old_table_rows == table_rows:
                soup = BeautifulSoup(self.driver.page_source, 'lxml')
                table = soup.find('table')
                table_rows = table.tbody.find_all('tr')
                time.sleep(2) # pause for 2 seconds cos it takes a while to load the next page of results
                if v % 5 == 0:
                    self.driver.find_elements('class name', 'Button')[1].click() # retries the 'next' button if it's been stuck on a page for too long
                v += 1
            print("Scheduled Update: Currently scraping page", int(arr[0])//50, '...') # some fancy math to keep track of progress

        return {"status_code": 0, "error_message": self.error_message, "success_message": self.success_response}
